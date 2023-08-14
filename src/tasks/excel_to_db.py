import re
from uuid import UUID

from openpyxl import load_workbook

from src.api.request_models.request_base import DishRequest, MenuRequest
from src.services.dishes_service import DishService
from src.services.menus_service import MenuService
from src.services.submenus_service import SubmenuService


class ExcelParser:
    """Class for parsing admin changes in excel file"""

    def __init__(
        self,
        menu_service: MenuService,
        sumbenu_service: SubmenuService,
        dish_service: DishService,
    ) -> None:
        self.__menu_service = menu_service
        self.__sumbenu_service = sumbenu_service
        self.__dish_service = dish_service

    async def parse_excel_and_load_to_db(self) -> None:
        wb = load_workbook('src/admin/Menu.xlsx')
        sheet = wb.active
        current_menu = None
        current_submenu = None
        current_dish = None
        excel_menu_uuids = set()
        excel_submenu_uuids = set()
        excel_dish_uuids = set()
        menu_uuids = set()
        submenu_uuids = set()
        dish_uuids = set()
        all_menus = await self.__menu_service.full_menus()
        uuid4_pattern = re.compile(
            r'^[a-f0-9]{8}-[a-f0-9]{4}-4[a-f0-9]{3}-[89aAbB][a-f0-9]{3}-[a-f0-9]{12}$',
            re.I,
        )

        updated_rows = []
        for menu in all_menus:
            menu_uuids.add(menu.id)
            for submenu in menu.submenus:
                submenu_uuids.add(f'{menu.id}:{submenu.id}')
                for dish in submenu.dishes:
                    dish_uuids.add(f'{menu.id}:{submenu.id}:{dish.id}')

        for row in sheet.iter_rows(values_only=True):
            updated_row = list(row)
            if isinstance(row[0], int):
                current_menu = await self.__menu_service.create_menu(
                    MenuRequest(title=row[1], description=row[2])
                )
                updated_row[0] = str(current_menu.id)
                current_submenu = None
            elif isinstance(row[0], str) and uuid4_pattern.match(row[0]):
                current_menu = await self.__menu_service.update_menu(
                    UUID(row[0]), MenuRequest(title=row[1], description=row[2])
                )
                excel_menu_uuids.add(current_menu.id)

            if current_menu and isinstance(row[1], int):
                current_submenu = await self.__sumbenu_service.create_submenu(
                    current_menu.id,
                    MenuRequest(title=row[2], description=row[3]),
                )
                updated_row[1] = str(current_submenu.id)
                current_dish = None
            elif current_menu and isinstance(row[1], str) and uuid4_pattern.match(row[1]):
                current_submenu = await self.__sumbenu_service.update_submenu(
                    UUID(row[1]), MenuRequest(title=row[2], description=row[3])
                )
                excel_submenu_uuids.add(
                    f'{current_menu.id}:{current_submenu.id}'
                )

            if current_menu and current_submenu and isinstance(row[2], int):
                current_dish = await self.__dish_service.create_dish(
                    current_menu.id,
                    current_submenu.id,
                    DishRequest(
                        title=row[3], description=row[4], price=row[5]
                    ),
                )
                updated_row[2] = str(current_dish.id)
            elif current_menu and current_submenu and isinstance(row[2], str) and uuid4_pattern.match(row[2]):
                current_dish = await self.__dish_service.update_dish(
                    current_menu.id,
                    current_submenu.id,
                    UUID(row[2]),
                    DishRequest(
                        title=row[3], description=row[4], price=row[5]
                    ),
                )
                excel_dish_uuids.add(
                    f'{current_menu.id}:{current_submenu.id}:{current_dish.id}'
                )

            updated_rows.append(tuple(updated_row))

        for idx, updated_row in enumerate(updated_rows):  # type: ignore
            for col_idx, value in enumerate(updated_row):
                sheet.cell(row=idx + 1, column=col_idx + 1, value=value)

        wb.save('src/admin/Menu.xlsx')

        deleted_dishes = dish_uuids - excel_dish_uuids
        if deleted_dishes:
            for uuid in deleted_dishes:
                menu_id, submenu_id, dish_id = uuid.split(':')
                await self.__dish_service.delete_dish(
                    UUID(menu_id), UUID(submenu_id), UUID(dish_id)
                )
        deleted_submenus = submenu_uuids - excel_submenu_uuids
        if deleted_submenus:
            for uuid in deleted_submenus:
                menu_id, submenu_id = uuid.split(':')
                await self.__sumbenu_service.delete_submenu(
                    UUID(menu_id), UUID(submenu_id)
                )
        deleted_menus = menu_uuids - excel_menu_uuids
        if deleted_menus:
            for uuid in deleted_menus:
                await self.__menu_service.delete_menu(UUID(uuid))
